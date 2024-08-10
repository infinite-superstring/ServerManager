import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from util.logger import Log
import os
import tempfile

class ColumnValidate(object):
    # 最小值，select有值时不生效
    min: int | None
    # 最大值，select有值时不生效
    max: int | None
    # 下拉选择
    select: list[str] | None
    # 错误标题
    error_title: str | None
    # 错误信息
    error_msg: str | None

    def __init__(self, min: int | None = None, max: int | None = None, select: list[str] | None = None,
                 error_title: str | None = None, error_msg: str | None = None):
        """
        :param min: 最小值
        :param max: 最大值
        :param select: 选项
        :param error_title: 错误标题
        :param error_msg: 错误提示信息
        """
        self.min = min
        self.max = max
        self.select = select
        self.error_title = error_title
        self.error_msg = error_msg

    def __str__(self):
        return (f"ColumnValidate(min={self.min}, max={self.max}, "
                f"select={self.select}, error_title={self.error_title}, "
                f"error_msg={self.error_msg})")


class ExcelColumn(object):
    """表格列"""
    # 列名
    column_name: str
    # 列数据类型
    col_type: str
    # 模板长度，输出模板文件时使用
    template_length: int
    # 默认值，表格中的默认值，为None不插入默认值
    default: any
    # 校验规则
    validate: ColumnValidate | None

    def __init__(self, column_name: str, col_type: str = 'str', template_length: int = 20, default: any = '',
                 validate: ColumnValidate | None = None):
        """
        :param column_name: 列名
        :param col_type: 列数据类型
        :param template_length: 列模板长度（减去表头）
        :param default: 列默认值
        :param validate: 校验规则
        """
        self.column_name = column_name
        self.col_type = col_type
        self.template_length = template_length
        self.default = default
        self.validate = validate

    def __str__(self):
        return (f"ExcelColumn(column_name='{self.column_name}', col_type='{self.col_type}', "
                f"template_length={self.template_length}, default={self.default}, "
                f"validate={self.validate})")


class ExcelRow(object):
    """表格行，用于存储表格数据信息，不包含表头"""
    # 表格数据
    data: list[any]
    # 校验错误，对应data列表，当错误时对应的bool为True
    error: list[bool]

    def __init__(self, data: list[any] = [], error: list[bool] = []):
        self.data = data
        self.error = error

    def __str__(self):
        return f"ExcelRow(\ndata={self.data}, \nerror={self.error}\n)"


class ExcelTable(object):
    """表格"""
    # 列
    cols: list[ExcelColumn]
    # 行
    rows: list[ExcelRow]

    def __init__(self, cols: list[ExcelColumn] = [], rows: list[ExcelRow] = []):
        self.cols = cols
        self.rows = rows

    def __str__(self):
        cols_str = ', \n'.join(str(col) for col in self.cols)
        rows_str = ', \n'.join(str(row) for row in self.rows)
        return f"ExcelTable(\ncols=[{cols_str}], \nrows=[{rows_str}]\n)"


class ExcelUtils(object):
    tables: dict[str:ExcelTable]
    # 创建加粗字体样式
    font = Font(bold=True)
    # 创建边框样式
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, tables=None):
        if tables is None:
            tables = {}
        self.tables = tables

    def _create_data_validation(self, data_type: str, rule: ColumnValidate) -> DataValidation:
        """
        创建数据校验规则
        :param data_type: 数据类型
        :param rule: 数据校验规则
        """
        dv = DataValidation()

        if rule.error_msg:
            dv.error = rule.error_msg
        else:
            dv.error = "数据值校验失败，请检查是否填写无误"
        if rule.error_title:
            dv.error_title = rule.error_title
        else:
            dv.error_title = "数据校验失败"
        dv.showErrorMessage = True

        if rule.min and rule.max:
            dv.formula1 = rule.min
            dv.formula2 = rule.max
            dv.operator = "between"
        else:
            if rule.min:
                dv.formula1 = rule.min
                dv.operator = "greaterThan"
            if rule.max:
                dv.formula1 = rule.max
                dv.operator = "lessThan"

        match data_type:
            case 'str':
                dv.type = "textLength"
                return dv
            case 'int':
                dv.type = "whole"
                return dv
            case 'float':
                dv.type = "decimal"
                return dv
            case 'bool':
                dv.type = "list"
                dv.allow_blank = True
                dv.formula1 = '"True,False"'
                dv.prompt = '请选择一个有效的选项'
                dv.promptTitle = '输入提示'
                return dv
            case 'select':
                dv.type = "list"
                dv.allow_blank = True
                dv.formula1 = f'"{",".join(rule.select)}"'
                dv.prompt = '请选择一个有效的选项'
                dv.promptTitle = '输入提示'
                return dv
            case _:
                Log.warning(f"Unknown data type: {data_type}")

        return dv

    def createExcelTemplate(self, save_path: str, return_wb_obj: bool=False):
        """
        创建Excel模板表格
        :param save_path: 输出文件路径
        """

        def _handle_table(work_book: openpyxl.Workbook, table_name: str, table_obj: ExcelTable):
            """
            处理工作表
            :param work_book: 工作簿对象
            :param table_name: 工作表名
            :param table_obj: 工作表对象
            """
            # 创建工作表
            work_table = work_book.create_sheet(title=table_name)
            # 创建表头
            for col_num, column in enumerate(table_obj.cols, start=1):
                cell = work_table.cell(row=1, column=col_num, value=column.col_type)
                cell.value = column.column_name
                cell.font = self.font
                cell.border = self.border
                col_letter = openpyxl.utils.get_column_letter(col_num)

                if column.validate:
                    data_rule = self._create_data_validation(column.col_type, column.validate)
                    work_table.add_data_validation(data_rule)
                    data_rule.add(f"{col_letter}2:{col_letter}{column.template_length+1}")

                for row in range(2, column.template_length + 1):
                    selected_cell = work_table[f"{col_letter}{row}"]
                    selected_cell.border = self.border  # 设置单元格边框
                    if column.default:
                        selected_cell.value = column.default

        wb = openpyxl.Workbook()
        for table_name, table in self.tables.items():
            _handle_table(wb, table_name, table)
        if not return_wb_obj:
            wb.save(save_path)
            return
        return wb


    def loadExcel(self, file_path):
        """
        读取Excel表格到对象并校验
        :param file_path: 要读取的文件
        """
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)

        for table_name, table_obj in self.tables.items():
            # 加载工作表
            if table_name not in wb.sheetnames:
                Log.error(f"工作表 {table_name} 不存在于文件 {file_path} 中")
                continue

            work_table = wb[table_name]
            rows = []

            # 读取每一行的数据
            for row_num in range(2, work_table.max_row + 1):
                data = []
                error = []
                for col_num, col_obj in enumerate(table_obj.cols, start=1):
                    cell_value = work_table.cell(row=row_num, column=col_num).value
                    data.append(cell_value)

                    # 校验数据
                    if col_obj.validate:
                        # 根据类型进行校验
                        if col_obj.col_type == "str":
                            if col_obj.validate.min and len(str(cell_value)) < col_obj.validate.min:
                                Log.error(f"第 {row_num} 行, 列 {col_num} 的字符串长度小于 {col_obj.validate.min}")
                                error.append(True)
                            elif col_obj.validate.max and len(str(cell_value)) > col_obj.validate.max:
                                Log.error(f"第 {row_num} 行, 列 {col_num} 的字符串长度大于 {col_obj.validate.max}")
                                error.append(True)
                            else:
                                error.append(False)
                        elif col_obj.col_type in ["int", "float"]:
                            try:
                                value = float(cell_value) if col_obj.col_type == "float" else int(cell_value)
                                if col_obj.validate.min and value < col_obj.validate.min:
                                    Log.error(f"第 {row_num} 行, 列 {col_num} 的值小于 {col_obj.validate.min}")
                                    error.append(True)
                                elif col_obj.validate.max and value > col_obj.validate.max:
                                    Log.error(f"第 {row_num} 行, 列 {col_num} 的值大于 {col_obj.validate.max}")
                                    error.append(True)
                                else:
                                    error.append(False)
                            except (ValueError, TypeError):
                                Log.error(f"第 {row_num} 行, 列 {col_num} 的值不是有效的数字")
                                error.append(True)
                        elif col_obj.col_type == "bool":
                            if cell_value not in ["True", "False", None]:
                                Log.error(f"第 {row_num} 行, 列 {col_num} 的值不是有效的布尔值")
                                error.append(True)
                            else:
                                error.append(False)
                        elif col_obj.col_type == "select":
                            if cell_value not in col_obj.validate.select:
                                Log.error(f"第 {row_num} 行, 列 {col_num} 的值不在有效选项中")
                                error.append(True)
                            else:
                                error.append(False)
                        else:
                            error.append(False)  # 默认无错误
                    else:
                        error.append(False)  # 没有验证规则则默认无错误
                rows.append(ExcelRow(data=data, error=error))
            table_obj.rows = rows

        wb.close()


if __name__ == '__main__':
    number = ColumnValidate(3, 10)
    string_length = ColumnValidate(3, 10)
    cols = [
        ExcelColumn("A1", default="111", validate=string_length),
        ExcelColumn("A2", 'int', validate=number),
        ExcelColumn("A3", 'float', validate=number),
        ExcelColumn("A4", 'bool', validate=ColumnValidate()),
        ExcelColumn("A5", 'select', validate=ColumnValidate(select=["测试1", "测试2", "测试3"])),
    ]
    eutils = ExcelUtils({"表1": ExcelTable(cols)})
    eutils.createExcelTemplate("file.xlsx")

    eutils.loadExcel("file.xlsx")
    for tab_name, table in eutils.tables.items():
        print(tab_name)
        print(table)
        print("="*10)